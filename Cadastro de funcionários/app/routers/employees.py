# app/routers/employees.py
from fastapi import APIRouter, Request, Depends, Form, HTTPException, UploadFile, File
from fastapi.responses import RedirectResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from starlette import status

# ORM e utilitários de banco de dados
from sqlalchemy.orm import Session
from sqlalchemy import select
from pathlib import Path

# Manipulação de arquivos
from io import BytesIO
from openpyxl import load_workbook

# Importações do projeto
from app.database import get_db, engine, Base
from app.models import Employee # Alterado

# Ajuste os helpers se necessário, mas os filtros parecem genéricos
from app.helpers import format_brl_price, format_brl_date, parse_brl_price

# Cria o roteador de funcionários
router = APIRouter()

# Cria as tabelas (incluindo a de funcionários)
Base.metadata.create_all(bind=engine)

# Configuração de diretório de templates
TEMPLATES_DIR = Path(__file__).resolve().parent.parent / "templates"
templates = Jinja2Templates(directory=str(TEMPLATES_DIR))
templates.env.filters["brl_price"] = format_brl_price
templates.env.filters["brl_date"] = format_brl_date


# ------------------------------------------------------------------
# Redireciona a raiz "/" para "/employees"
# ------------------------------------------------------------------
@router.get("/", include_in_schema=False)
def root_redirect():
    return RedirectResponse(url="/employees", status_code=status.HTTP_302_FOUND)


# ------------------------------------------------------------------
# LISTAR FUNCIONÁRIOS (GET)
# ------------------------------------------------------------------
@router.get("/employees")
def list_employees(request: Request, db: Session = Depends(get_db)):
    employees = db.scalars(select(Employee).order_by(Employee.id.desc())).all()
    return templates.TemplateResponse(
        "employees/index.html", # Alterado
        {"request": request, "employees": employees} # Alterado
    )


# ------------------------------------------------------------------
# FORMULÁRIO DE NOVO FUNCIONÁRIO (GET)
# ------------------------------------------------------------------
@router.get("/employees/new")
def new_employee_form(request: Request):
    return templates.TemplateResponse(
        "employees/new.html", # Alterado
        {
            "request": request,
            "action": "/employees",         # Alterado
            "method_override": "POST",
            "employee": None                # Alterado
        }
    )


# ------------------------------------------------------------------
# FORMULÁRIO DE EDIÇÃO DE FUNCIONÁRIO (GET)
# ------------------------------------------------------------------
@router.get("/employees/{employee_id}/edit")
def edit_employee_form(employee_id: int, request: Request, db: Session = Depends(get_db)):
    employee = db.get(Employee, employee_id) # Alterado
    if not employee:
        raise HTTPException(status_code=404, detail="Funcionário não encontrado")

    return templates.TemplateResponse(
        "employees/edit.html", # Alterado
        {
            "request": request,
            "action": f"/employees/{employee.id}", # Alterado
            "method_override": "PUT",
            "employee": employee # Alterado
        }
    )


# ------------------------------------------------------------------
# CRIAÇÃO DE FUNCIONÁRIO (POST)
# ------------------------------------------------------------------
@router.post("/employees")
def create_employee(
    name: str = Form(...),
    email: str = Form(...),
    phone: str = Form(None),
    position: str = Form(None),
    department: str = Form(None),
    salary: float = Form(0.0),
    db: Session = Depends(get_db),
):
    if not name.strip() or not email.strip():
        raise HTTPException(status_code=400, detail="Nome e Email são obrigatórios")

    # Verifica se já existe um funcionário com o mesmo email
    if db.scalar(select(Employee).where(Employee.email == email)):
        raise HTTPException(status_code=400, detail="Email já cadastrado")

    # Cria o objeto e salva no banco
    db.add(Employee(
        name=name.strip(), 
        email=email.strip(),
        phone=phone.strip() if phone else None,
        position=position.strip() if position else None,
        department=department.strip() if department else None,
        salary=salary
    ))
    db.commit()

    return RedirectResponse(url="/employees", status_code=status.HTTP_303_SEE_OTHER)


# ------------------------------------------------------------------
# ATUALIZAÇÃO DE FUNCIONÁRIO (PUT)
# ------------------------------------------------------------------
@router.put("/employees/{employee_id}")
def update_employee(
    employee_id: int,
    name: str = Form(...),
    email: str = Form(...),
    phone: str = Form(None),
    position: str = Form(None),
    department: str = Form(None),
    salary: float = Form(0.0),
    db: Session = Depends(get_db),
):
    employee = db.get(Employee, employee_id) # Alterado
    if not employee:
        raise HTTPException(status_code=404, detail="Funcionário não encontrado")

    # Verifica se o Email já pertence a outro funcionário
    if db.scalar(select(Employee).where(Employee.email == email, Employee.id != employee_id)):
        raise HTTPException(status_code=400, detail="Email já cadastrado para outro funcionário")

    # Atualiza os campos
    employee.name = name.strip()
    employee.email = email.strip()
    employee.phone = phone.strip() if phone else None
    employee.position = position.strip() if position else None
    employee.department = department.strip() if department else None
    employee.salary = salary

    db.add(employee)
    db.commit()

    return JSONResponse(status_code=status.HTTP_204_NO_CONTENT, content=None)


# ------------------------------------------------------------------
# EXCLUSÃO DE FUNCIONÁRIO (DELETE)
# ------------------------------------------------------------------
@router.delete("/employees/{employee_id}")
def delete_employee(employee_id: int, db: Session = Depends(get_db)):
    employee = db.get(Employee, employee_id) # Alterado
    if not employee:
        raise HTTPException(status_code=404, detail="Funcionário não encontrado")

    db.delete(employee)
    db.commit()

    return JSONResponse(status_code=status.HTTP_204_NO_CONTENT, content=None)


# ------------------------------------------------------------------
# IMPORTAR DADOS DO EXCEL
# ------------------------------------------------------------------

@router.get("/employees/import")
def import_employees_form(request: Request):
    return templates.TemplateResponse(
        "employees/import.html", # Alterado
        {"request": request, "report": None}
    )


@router.post("/employees/import")
async def import_employees(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """
    Recebe um arquivo .xlsx com colunas obrigatórias:
    Nome, Email, Salário, Cargo, Departamento, Telefone
    """
    allowed = {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel", "application/octet-stream",
    }
    if file.content_type not in allowed:
        raise HTTPException(status_code=400, detail="Arquivo inválido. Envie um .xlsx")

    content = await file.read()
    try:
        wb = load_workbook(BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Não foi possível ler o arquivo .xlsx")

    ws = wb.active
    if ws.max_row < 2:
        return templates.TemplateResponse(
            "employees/import.html",
            {"request": request, "report": {"imported": 0, "skipped": 0, "errors": [{"row": 1, "error": "Planilha vazia"}]}}
        )

    # Validação de cabeçalhos
    headers = [str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
    expected_headers = ["Nome", "Email", "Salário", "Cargo", "Departamento", "Telefone"]

    if headers[:6] != expected_headers:
        raise HTTPException(
            status_code=400,
            detail=f"Cabeçalhos inválidos. Esperado: {', '.join(expected_headers)}"
        )

    header_map = {"name": 1, "email": 2, "salary": 3, "position": 4, "department": 5, "phone": 6}

    # Preparar validações
    existing_emails = set(db.scalars(select(Employee.email)).all())
    seen_emails_in_file = set()

    imported = 0
    skipped = 0
    errors = []
    new_employees = []

    # Iterar linhas de dados (a partir da linha 2)
    for row in range(2, ws.max_row + 1):
        try:
            raw_name = ws.cell(row=row, column=header_map["name"]).value
            raw_email = ws.cell(row=row, column=header_map["email"]).value
            raw_salary = ws.cell(row=row, column=header_map["salary"]).value
            raw_position = ws.cell(row=row, column=header_map["position"]).value
            raw_department = ws.cell(row=row, column=header_map["department"]).value
            raw_phone = ws.cell(row=row, column=header_map["phone"]).value

            name = (str(raw_name).strip() if raw_name else "")
            email = (str(raw_email).strip() if raw_email else "")
            phone = (str(raw_phone).strip() if raw_phone else None)
            position = (str(raw_position).strip() if raw_position else None)
            department = (str(raw_department).strip() if raw_department else None)

            if not name: raise ValueError("Nome vazio")
            if not email: raise ValueError("Email vazio")

            salary = parse_brl_price(raw_salary)

            if email in existing_emails:
                raise ValueError("Email já cadastrado no banco")
            if email in seen_emails_in_file:
                raise ValueError("Email duplicado no arquivo")

            new_employees.append(Employee(
                name=name, email=email, salary=salary, 
                position=position, department=department, phone=phone
            ))
            seen_emails_in_file.add(email)
            imported += 1

        except Exception as e:
            skipped += 1
            errors.append({"row": row, "error": str(e)})

    # Inserir válidos
    if new_employees:
        db.add_all(new_employees)
        db.commit()

    report = {"imported": imported, "skipped": skipped, "errors": errors}
    return templates.TemplateResponse(
        "employees/import.html",
        {"request": request, "report": report}
    )
    
# ------------------------------------------------------------------
# DETALHE DE FUNCIONÁRIO (GET)
# ------------------------------------------------------------------
@router.get("/employees/{employee_id}")
def get_employee(employee_id: int, request: Request, db: Session = Depends(get_db)):
    employee = db.get(Employee, employee_id) # Alterado
    if not employee:
        raise HTTPException(status_code=404, detail="Funcionário não encontrado")

    return templates.TemplateResponse(
        "employees/show.html", # Alterado
        {"request": request, "employee": employee} # Alterado
    )